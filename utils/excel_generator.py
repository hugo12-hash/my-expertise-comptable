"""
Génération du fichier Excel comptable normalisé.
Pour chaque transaction, génère 2 lignes (partie double) :
- Ligne 1 : compte banque (512)
- Ligne 2 : compte contrepartie (471, sens inversé)
"""
import io
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = ['DATE', 'PIECE', 'COMPTE', 'LIB', 'DEBIT', 'CREDIT']


def build_comptable_rows(
    transactions: List[Dict],
    compte_banque: str = "51200000",
    compte_contrepartie: str = "47100000",
) -> List[Dict]:
    """
    Transforme une liste de transactions en lignes comptables avec partie double.
    
    Pour chaque transaction :
    - Si DÉBIT (sortie d'argent) :
        * Ligne 1 : 471 débit, 512 crédit (logique : argent sort du 512, en attente sur 471)
        * En réalité dans un schéma simple : 512 crédit, 471 débit
    - Si CRÉDIT (entrée d'argent) :
        * Ligne 1 : 512 débit, 471 crédit
    
    Logique standard expertise comptable :
    - Le 512 reflète exactement le mouvement bancaire (débit/crédit comme la banque)
    - Le 471 fait la contrepartie en sens inverse (compte d'attente)
    """
    rows = []
    
    for t in transactions:
        date = t.get('date', '')
        libelle = t.get('libelle', '')
        debit = t.get('debit')
        credit = t.get('credit')
        
        # Ignorer les transactions sans montant
        if not debit and not credit:
            continue
        
        if credit:  # Entrée d'argent (la banque crédite le compte)
            # Ligne 1 : Banque 512 au débit (l'argent rentre)
            rows.append({
                'DATE': date,
                'PIECE': '',
                'COMPTE': compte_banque,
                'LIB': libelle,
                'DEBIT': credit,
                'CREDIT': '',
            })
            # Ligne 2 : Contrepartie 471 au crédit
            rows.append({
                'DATE': date,
                'PIECE': '',
                'COMPTE': compte_contrepartie,
                'LIB': libelle,
                'DEBIT': '',
                'CREDIT': credit,
            })
        elif debit:  # Sortie d'argent (la banque débite le compte)
            # Ligne 1 : Banque 512 au crédit (l'argent sort)
            rows.append({
                'DATE': date,
                'PIECE': '',
                'COMPTE': compte_banque,
                'LIB': libelle,
                'DEBIT': '',
                'CREDIT': debit,
            })
            # Ligne 2 : Contrepartie 471 au débit
            rows.append({
                'DATE': date,
                'PIECE': '',
                'COMPTE': compte_contrepartie,
                'LIB': libelle,
                'DEBIT': debit,
                'CREDIT': '',
            })
    
    return rows


def generate_excel(
    transactions: List[Dict],
    compte_banque: str = "51200000",
    compte_contrepartie: str = "47100000",
    filename_label: str = "Releve",
) -> io.BytesIO:
    """
    Génère un fichier Excel formaté à partir des transactions.
    Retourne un BytesIO prêt à être téléchargé.
    """
    rows = build_comptable_rows(transactions, compte_banque, compte_contrepartie)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ecritures"
    
    # ----- En-têtes -----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D85A30", end_color="D85A30", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='E8E2DC'),
        right=Side(style='thin', color='E8E2DC'),
        top=Side(style='thin', color='E8E2DC'),
        bottom=Side(style='thin', color='E8E2DC'),
    )
    
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # ----- Données -----
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = thin_border
            # Format nombre pour débit/crédit
            if col_name in ('DEBIT', 'CREDIT') and row[col_name] != '':
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            # Centrer date, pièce, compte
            elif col_name in ('DATE', 'PIECE', 'COMPTE'):
                cell.alignment = Alignment(horizontal="center")
    
    # ----- Largeurs des colonnes -----
    widths = {'DATE': 12, 'PIECE': 10, 'COMPTE': 12, 'LIB': 50, 'DEBIT': 14, 'CREDIT': 14}
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]
    
    # ----- Figer la première ligne -----
    ws.freeze_panes = 'A2'
    
    # ----- Total en bas -----
    last_row = len(rows) + 1
    total_row = last_row + 2
    ws.cell(row=total_row, column=4, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="TOTAUX").alignment = Alignment(horizontal="right")
    
    total_debit_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_row})")
    total_debit_cell.font = Font(bold=True)
    total_debit_cell.number_format = '#,##0.00'
    
    total_credit_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{last_row})")
    total_credit_cell.font = Font(bold=True)
    total_credit_cell.number_format = '#,##0.00'
    
    # ----- Sauvegarde en mémoire -----
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_preview_dataframe(
    transactions: List[Dict],
    compte_banque: str = "51200000",
    compte_contrepartie: str = "47100000",
) -> pd.DataFrame:
    """Construit un DataFrame pandas pour l'aperçu dans Streamlit."""
    rows = build_comptable_rows(transactions, compte_banque, compte_contrepartie)
    df = pd.DataFrame(rows, columns=COLUMNS)
    return df
