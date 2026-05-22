"""
Traitement du CSV de relevé bancaire BANQUE ELECTRIP.
Transforme un export CSV brut en Excel comptable avec partie double.
"""
import io
import re
from typing import Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colonnes attendues dans le CSV source
COLONNES_REQUISES = ['valeur_date', 'nom_contrepartie', 'libelle', 'montant_transaction']

# Colonnes finales de la feuille "Données brutes"
COLONNES_DONNEES_BRUTES = [
    'valeur_date', 'nom_contrepartie', 'libelle', 'Lib',
    'Compte', 'Debit', 'Credit', 'Contrepartie'
]

# Colonnes finales de la feuille "Écritures comptables"
COLONNES_ECRITURES = ['DATE', 'PIECE', 'COMPTE', 'LIB', 'DEBIT', 'CREDIT']


def lire_csv(csv_bytes: bytes) -> pd.DataFrame:
    """Lit le CSV en gérant les encodages courants."""
    # Essayer plusieurs encodages dans l'ordre
    for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
        try:
            df = pd.read_csv(io.BytesIO(csv_bytes), encoding=encoding, dtype=str)
            return df
        except UnicodeDecodeError:
            continue
    # En dernier recours, on force avec gestion d'erreurs
    return pd.read_csv(io.BytesIO(csv_bytes), encoding='utf-8', errors='replace', dtype=str)


def normaliser_date(date_str: str) -> str:
    """
    Convertit '22-05-2026 12:43:04' en '22/05/2026'.
    Gère aussi le format JJ/MM/AAAA déjà normalisé.
    """
    if pd.isna(date_str) or not date_str:
        return ''
    s = str(date_str).strip()
    # Extraire la partie date (avant l'espace) si présente
    if ' ' in s:
        s = s.split(' ')[0]
    # Remplacer tirets par slashes
    s = s.replace('-', '/').replace('.', '/')
    # Vérifier qu'on a bien JJ/MM/AAAA
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', s)
    if not m:
        return s  # On retourne tel quel si pas reconnu
    d, mo, y = m.groups()
    if len(y) == 2:
        y = '20' + y if int(y) < 50 else '19' + y
    return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"


def parser_montant(montant_str) -> float:
    """Convertit un montant string ('-13.5', '1521.17') en float."""
    if pd.isna(montant_str) or montant_str == '':
        return 0.0
    s = str(montant_str).strip()
    # Gérer les deux séparateurs : virgule et point
    s = s.replace(',', '.')
    # Retirer les espaces (séparateurs de milliers éventuels)
    s = s.replace(' ', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def construire_donnees_brutes(
    df_source: pd.DataFrame,
    compte_banque: str,
    compte_contrepartie: str,
) -> pd.DataFrame:
    """
    Étape 1 : Construit le DataFrame de la feuille "Données brutes"
    avec toutes les transformations et les contreparties.
    """
    # Vérifier les colonnes requises
    colonnes_manquantes = [c for c in COLONNES_REQUISES if c not in df_source.columns]
    if colonnes_manquantes:
        raise ValueError(
            f"Colonnes manquantes dans le CSV : {', '.join(colonnes_manquantes)}. "
            f"Colonnes attendues : {', '.join(COLONNES_REQUISES)}"
        )

    # Garder uniquement les colonnes utiles
    df = df_source[COLONNES_REQUISES].copy()

    # Normaliser les dates au format JJ/MM/AAAA
    df['valeur_date'] = df['valeur_date'].apply(normaliser_date)

    # Nettoyer les NaN
    df['nom_contrepartie'] = df['nom_contrepartie'].fillna('').astype(str).str.strip()
    df['libelle'] = df['libelle'].fillna('').astype(str).str.strip()

    # Créer la colonne Lib = nom_contrepartie + " " + libelle
    df['Lib'] = (df['nom_contrepartie'] + ' ' + df['libelle']).str.strip()
    # Nettoyer les doubles espaces
    df['Lib'] = df['Lib'].apply(lambda x: re.sub(r'\s+', ' ', x))

    # Parser les montants en float
    df['_montant_float'] = df['montant_transaction'].apply(parser_montant)

    # Créer Debit et Credit selon le signe
    # Positif → Debit (entrée d'argent côté banque selon ta logique)
    # Négatif → Credit (sortie, en valeur absolue)
    df['Debit'] = df['_montant_float'].apply(lambda x: x if x > 0 else 0.0)
    df['Credit'] = df['_montant_float'].apply(lambda x: abs(x) if x < 0 else 0.0)

    # Compte banque pour toutes les lignes originales
    df['Compte'] = compte_banque
    df['Contrepartie'] = ''  # vide pour les lignes originales

    # Créer les contreparties (lignes dupliquées avec Débit/Crédit inversés)
    df_contre = df.copy()
    df_contre['Compte'] = compte_contrepartie
    df_contre['Debit'], df_contre['Credit'] = df_contre['Credit'].copy(), df_contre['Debit'].copy()
    df_contre['Contrepartie'] = 'X'

    # Fusionner les deux DataFrames en alternant (1 originale, 1 contrepartie, ...)
    # On crée une colonne d'ordre pour bien intercaler
    df['_ordre'] = range(0, len(df) * 2, 2)
    df_contre['_ordre'] = range(1, len(df_contre) * 2 + 1, 2)

    df_final = pd.concat([df, df_contre], ignore_index=True)
    df_final = df_final.sort_values('_ordre').reset_index(drop=True)

    # Nettoyer les colonnes de travail
    df_final = df_final.drop(columns=['_montant_float', '_ordre'])

    # Réordonner les colonnes
    df_final = df_final[COLONNES_DONNEES_BRUTES]

    return df_final


def construire_ecritures(df_brutes: pd.DataFrame) -> pd.DataFrame:
    """
    Étape 2 : Construit le DataFrame de la feuille "Écritures comptables"
    au format DATE | PIECE | COMPTE | LIB | DEBIT | CREDIT
    """
    df_ecr = pd.DataFrame({
        'DATE': df_brutes['valeur_date'],
        'PIECE': '',  # vide
        'COMPTE': df_brutes['Compte'],
        'LIB': df_brutes['Lib'],
        'DEBIT': df_brutes['Debit'],
        'CREDIT': df_brutes['Credit'],
    })
    return df_ecr


def formater_montant_excel(val: float) -> str:
    """Formate un montant float en string avec virgule (ex: 13.5 → '13,50')."""
    if val == 0 or pd.isna(val):
        return ''
    return f"{val:.2f}".replace('.', ',')


def generer_excel(
    df_source: pd.DataFrame,
    compte_banque: str = "5120000",
    compte_contrepartie: str = "4710000",
) -> Tuple[io.BytesIO, pd.DataFrame, pd.DataFrame]:
    """
    Génère un fichier Excel à 2 feuilles : Données brutes + Écritures comptables.
    Retourne (BytesIO, df_brutes, df_ecritures) pour pouvoir aussi afficher des aperçus.
    """
    # Construire les deux DataFrames
    df_brutes = construire_donnees_brutes(df_source, compte_banque, compte_contrepartie)
    df_ecritures = construire_ecritures(df_brutes)

    # Créer le workbook
    wb = Workbook()

    # ---------- Feuille 1 : Données brutes ----------
    ws1 = wb.active
    ws1.title = "Donnees brutes"
    _ecrire_feuille_brutes(ws1, df_brutes)

    # ---------- Feuille 2 : Écritures comptables ----------
    ws2 = wb.create_sheet(title="Ecritures comptables")
    _ecrire_feuille_ecritures(ws2, df_ecritures)

    # Sauvegarder en mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, df_brutes, df_ecritures


def _appliquer_style_entete(cell, couleur_hex="D85A30"):
    """Applique le style d'en-tête (fond orange MY Expertise)."""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=couleur_hex, end_color=couleur_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin', color='E8E2DC'),
        right=Side(style='thin', color='E8E2DC'),
        top=Side(style='thin', color='E8E2DC'),
        bottom=Side(style='thin', color='E8E2DC'),
    )


def _ecrire_feuille_brutes(ws, df: pd.DataFrame):
    """Écrit la feuille 'Données brutes' avec mise en forme."""
    # En-têtes
    for col_idx, col_name in enumerate(COLONNES_DONNEES_BRUTES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _appliquer_style_entete(cell)

    # Données
    thin_border = Border(
        left=Side(style='thin', color='E8E2DC'),
        right=Side(style='thin', color='E8E2DC'),
        top=Side(style='thin', color='E8E2DC'),
        bottom=Side(style='thin', color='E8E2DC'),
    )

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(COLONNES_DONNEES_BRUTES, start=1):
            val = row[col_name]

            # Formatage spécial pour Debit/Credit
            if col_name in ('Debit', 'Credit'):
                if val == 0 or val == '' or pd.isna(val):
                    val_to_write = ''
                else:
                    val_to_write = float(val)
            else:
                val_to_write = val if not pd.isna(val) else ''

            cell = ws.cell(row=row_idx, column=col_idx, value=val_to_write)
            cell.border = thin_border

            if col_name in ('Debit', 'Credit'):
                if val_to_write != '':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            elif col_name in ('valeur_date', 'Compte', 'Contrepartie'):
                cell.alignment = Alignment(horizontal="center")

            # Surligner les lignes de contrepartie (fond beige clair)
            if col_name == 'Contrepartie' and val_to_write == 'X':
                # On marque toute la ligne
                for c_idx in range(1, len(COLONNES_DONNEES_BRUTES) + 1):
                    target_cell = ws.cell(row=row_idx, column=c_idx)
                    target_cell.fill = PatternFill(
                        start_color="FAF6F2", end_color="FAF6F2", fill_type="solid"
                    )

    # Largeurs
    widths = {
        'valeur_date': 12, 'nom_contrepartie': 28, 'libelle': 45, 'Lib': 50,
        'Compte': 12, 'Debit': 14, 'Credit': 14, 'Contrepartie': 13
    }
    for col_idx, col_name in enumerate(COLONNES_DONNEES_BRUTES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]

    ws.freeze_panes = 'A2'


def _ecrire_feuille_ecritures(ws, df: pd.DataFrame):
    """Écrit la feuille 'Écritures comptables' au format compta."""
    # En-têtes
    for col_idx, col_name in enumerate(COLONNES_ECRITURES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _appliquer_style_entete(cell)

    thin_border = Border(
        left=Side(style='thin', color='E8E2DC'),
        right=Side(style='thin', color='E8E2DC'),
        top=Side(style='thin', color='E8E2DC'),
        bottom=Side(style='thin', color='E8E2DC'),
    )

    # Mapping df → colonnes Excel
    col_map = {'DATE': 'DATE', 'PIECE': 'PIECE', 'COMPTE': 'COMPTE',
               'LIB': 'LIB', 'DEBIT': 'DEBIT', 'CREDIT': 'CREDIT'}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(COLONNES_ECRITURES, start=1):
            val = row[col_map[col_name]]

            if col_name in ('DEBIT', 'CREDIT'):
                if val == 0 or val == '' or pd.isna(val):
                    val_to_write = ''
                else:
                    val_to_write = float(val)
            else:
                val_to_write = val if not pd.isna(val) else ''

            cell = ws.cell(row=row_idx, column=col_idx, value=val_to_write)
            cell.border = thin_border

            if col_name in ('DEBIT', 'CREDIT'):
                if val_to_write != '':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            elif col_name in ('DATE', 'PIECE', 'COMPTE'):
                cell.alignment = Alignment(horizontal="center")

    # Largeurs
    widths = {'DATE': 12, 'PIECE': 10, 'COMPTE': 12, 'LIB': 60, 'DEBIT': 14, 'CREDIT': 14}
    for col_idx, col_name in enumerate(COLONNES_ECRITURES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]

    # Total en bas
    last_row = len(df) + 1
    total_row = last_row + 2
    cell_label = ws.cell(row=total_row, column=4, value="TOTAUX")
    cell_label.font = Font(bold=True)
    cell_label.alignment = Alignment(horizontal="right")

    total_debit_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_row})")
    total_debit_cell.font = Font(bold=True)
    total_debit_cell.number_format = '#,##0.00'
    total_debit_cell.alignment = Alignment(horizontal="right")

    total_credit_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{last_row})")
    total_credit_cell.font = Font(bold=True)
    total_credit_cell.number_format = '#,##0.00'
    total_credit_cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = 'A2'
